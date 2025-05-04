# import openpyxl module
import openpyxl
from openpyxl import Workbook

# Give the location of the file
path = "Sites Management Website VS.xlsx"

# To open the workbook
# workbook object is created
try:
    wb_obj = openpyxl.load_workbook(path)
except FileNotFoundError:
    print(f"File {path} not found. Please check the file path and try again.")

sheet_obj = wb_obj.active

#قراءة كم سطر و كم عامود
# row = sheet_obj.max_row
# column = sheet_obj.max_column

#إجمالي عدد الأسطر و الأعمدة
# print("Total Rows:", row)
# print("Total Columns:", column)

#عم اطبع قيمة اول عامود
# print("\nValue of first column")
# for i in range(1, row + 1):
#     cell_obj = sheet_obj.cell(row=i, column=1)
#     print(cell_obj.value)


#عم اطبع قيمة أول سطر
# print("\nValue of first row")
# for i in range(1, column + 1):
#     cell_obj = sheet_obj.cell(row=1, column=i)
#     print(cell_obj.value, end=" ")
 
#عم اطبع السيلات ضمن مجال معين   
cell_obj = sheet_obj['A1': 'B7']
print("\nValue of  :")
for cell1, cell2 in cell_obj:
    print(cell1.value, cell2.value)
    
#عم كريت فايل اكسيل جديد
workbook = Workbook()

workbook.save(filename="sample.xlsx")

c1 = sheet_obj.cell(row=1, column=1)

# writing values to cells
c1.value = "Hello"

c2 = sheet_obj.cell(row=1, column=2)
c2.value = "World"

c3 = sheet_obj['A2']
c3.value = "Welcome"

# B2 means column = 2 & row = 2.
c4 = sheet_obj['B2']
c4.value = "Everyone"

#لحتى سيف عل الفايل
wb_obj.save("sample.xlsx")